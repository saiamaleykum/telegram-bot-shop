import os
from openpyxl import Workbook, load_workbook


def is_sheet_empty(worksheet):
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                return False
    return True


def order_to_excel(order_id, user_id, dt, address, items_in_cart):
    log_dir = '/app/excel'
    if not os.path.exists(log_dir):
        os.makedirs(log_dir)

    filename = os.path.join(log_dir, 'orders.xlsx')

    try:
        wb = load_workbook(filename)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = "Orders"

    if is_sheet_empty(ws):
        headers = ["Order ID", "User ID", "Time Created", "Status", "Address", "Item ID", "Title", "Quantity"]
        ws.append(headers)

    for i in range(len(items_in_cart)):
        if i == 0:
            ws.append([
                order_id,
                user_id,
                dt,
                "paid",
                address,
                items_in_cart[0][4],
                items_in_cart[0][1],
                items_in_cart[0][0]
            ])
        else:
            ws.append([
                "",
                "",
                "",
                "",
                "",
                items_in_cart[i][4],
                items_in_cart[i][1],
                items_in_cart[i][0]
            ])



    wb.save(filename)
